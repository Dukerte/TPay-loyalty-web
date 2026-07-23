#!/usr/bin/env python3
"""
TPay Loyalty – Excel → data.json / data2.json + draw_data.json converter
=========================================================================
Хэрэглээ / Usage:
  1-р сарын аян (Шинэ зээлийн):
    python3 convert.py month1_data.xlsx

  2-р сарын аян (Эргэн төлөлтийн):
    python3 convert.py month2_data.xlsx --month 2

  3-р сарын аян (Давтан зээлийн):
    python3 convert.py month3_data.xlsx --month 3

Excel файлын формат (олон хэлбэр автоматаар илэрнэ):

  [1] Олон хуудас (Data.xlsx — Эргэн төлөлтийн аян)
      Хуудас "Зээл сунгалт": 1-р мөр = толгой, 2-р мөрөөс өгөгдөл
        Баганууд: (хоосон) | Регистр | Овог | Нэр | Огноо | Утас
        1 сугалааны эрх
      Хуудас "Хаасан зээл":  2-р мөр = толгой, 3-р мөрөөс өгөгдөл
        Баганууд: (хоосон) | Регистр | Овог | Нэр | Огноо | Утас
        2 сугалааны эрх

  [2] TPay экспорт (1 хуудас, толгойн 1-р мөр нь "Төрөл" агуулна)
        Төрөл | Регистр | Овог | Нэр | Утас
        Сунгалт=1эрх, Хаасан/Хаах=2эрх

  [3] 5 багана (гараар бэлдсэн)
        Нэр | Овог | Утас | Киоск | Огноо  → 1эрх

  [4] 2 багана (хуучин)
        Утас | Тикет тоо

Гаралт:
  --month 1 (default) → data.json  + draw_data.json
  --month 2           → data2.json + draw_data2.json
  --month 3           → data3.json + draw_data3.json
"""

import sys
import json
import os
import re
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl олдсонгүй. Автоматаар суулгаж байна...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    print("✅ openpyxl суулгагдлаа.\n")


def clean_phone(raw):
    if raw is None:
        return None
    phone = str(raw).strip().replace(" ", "").replace("-", "")
    # Remove leading +976 or 976
    if phone.startswith("+976"):
        phone = phone[4:]
    elif phone.startswith("976") and len(phone) == 11:
        phone = phone[3:]
    if not phone.isdigit() or len(phone) != 8:
        return None
    return phone


def fmt_date(raw):
    if raw is None:
        return ""
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    # try common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return s[:10]


def detect_format(headers):
    """
    'tpay' — TPay системийн шууд экспорт: Төрөл | Регистр | Овог | Нэр | Утас
    'new'  — гараар бэлдсэн 5 багана:     Нэр | Овог | Утас | Киоск | Огноо
    'old'  — хуучин 2 багана:              Утас | Тикет тоо
    """
    h = [str(x).strip().lower() if x else "" for x in headers]
    if any("төрөл" in x for x in h):
        return "tpay"
    if any("нэр" in x or "овог" in x for x in h):
        return "new"
    return "old"


def write_output(entries, phone_counts, skipped, month, base_dir):
    """Write dataN.json and draw_dataN.json."""
    suffix = "" if month == 1 else str(month)
    total_tickets = len(entries)
    unique_phones = len(phone_counts)
    today = str(date.today())

    data_json = {
        "_updated": today,
        "_note": f"convert.py --month {month} ашиглан үүсгэсэн.",
    }
    data_json.update(phone_counts)

    data_json_path = os.path.join(base_dir, f"data{suffix}.json")
    with open(data_json_path, "w", encoding="utf-8") as f:
        json.dump(data_json, f, ensure_ascii=False, indent=2)

    draw_json = {
        "_updated": today,
        "_total_tickets": total_tickets,
        "_unique_phones": unique_phones,
        "entries": entries
    }
    draw_json_path = os.path.join(base_dir, f"draw_data{suffix}.json")
    with open(draw_json_path, "w", encoding="utf-8") as f:
        json.dump(draw_json, f, ensure_ascii=False, indent=2)

    print(f"✅  Амжилттай үүслээ")
    print(f"   Нийт тасалбар:     {total_tickets}")
    print(f"   Өвөрмөц харилцагч: {unique_phones}")
    print(f"   Алгасав:           {skipped}")
    print(f"   {data_json_path}")
    print(f"   {draw_json_path}")
    print()


def find_header_row(rows):
    """
    Return (header_row_idx, phone_col_idx) by scanning for 'утас' in any row.
    Returns (-1, -1) if not found.
    """
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell and "утас" in str(cell).strip().lower():
                return i, j
    return -1, -1


def process_sheet(ws, tickets_per_row, label):
    """
    Generic sheet processor: auto-detects header row and phone column.
    Returns (entries, phone_counts, ok_count, skip_count).
    """
    rows = [r for r in ws.iter_rows(values_only=True)]
    hdr_idx, phone_col = find_header_row(rows)
    if hdr_idx == -1:
        print(f"   ⚠️  [{label}] Толгой мөр олдсонгүй, алгасав.")
        return [], {}, 0, 0

    entries = []
    phone_counts = {}
    ok = skip = 0

    # detect name/surname columns from header
    hdr = [str(c).strip().lower() if c else "" for c in rows[hdr_idx]]
    # овог = surname, нэр = name (last occurrence if multiple)
    surname_col = next((j for j, h in enumerate(hdr) if "овог" in h), None)
    name_col    = next((j for j, h in enumerate(hdr) if "нэр" in h and "регистр" not in h), None)

    for row in rows[hdr_idx + 1:]:
        if not any(row):
            continue
        phone = clean_phone(row[phone_col]) if phone_col < len(row) else None
        if phone is None:
            skip += 1
            continue
        surname = str(row[surname_col]).strip() if surname_col is not None and row[surname_col] else ""
        name    = str(row[name_col]).strip()    if name_col    is not None and row[name_col]    else ""
        for _ in range(tickets_per_row):
            entries.append({"phone": phone, "name": name, "surname": surname, "kiosk": "", "date": ""})
        phone_counts[phone] = phone_counts.get(phone, 0) + tickets_per_row
        ok += 1

    return entries, phone_counts, ok, skip


def find_sheet(wb, keywords):
    """Return first sheet whose name matches any keyword (case-insensitive)."""
    for name in wb.sheetnames:
        n = name.strip().lower()
        if any(kw in n for kw in keywords):
            return wb[name], name
    return None, None


def process_multi_sheet(wb, month, base_dir, merge=False):
    """
    Олон хуудас формат — сунгалт + хаасан зээл хуудас автоматаар илэрнэ.
    Утасны багана болон толгой мөрийг динамикаар олно.
    --merge флагтай бол одоогийн dataN.json дээр нэмнэ.
    """
    sun_ws, sun_name = find_sheet(wb, ["сунгалт", "sungalt", "зээл сунгалт"])
    haa_ws, haa_name = find_sheet(wb, ["хаасан", "haasan"])

    if sun_ws is None and haa_ws is None:
        print("   ❌ Сунгалт/Хаасан хуудас олдсонгүй.")
        wb.close()
        return

    print(f"   Формат илэрлээ: Олон хуудас ({sun_name or '—'} + {haa_name or '—'})")

    entries = []
    phone_counts = {}
    skipped = 0

    if sun_ws:
        e, pc, ok, skip = process_sheet(sun_ws, 1, sun_name)
        entries += e
        for p, v in pc.items():
            phone_counts[p] = phone_counts.get(p, 0) + v
        skipped += skip
        print(f"   {sun_name}: {ok} мөр (1 эрх) · {skip} алгасав")

    if haa_ws:
        e, pc, ok, skip = process_sheet(haa_ws, 2, haa_name)
        entries += e
        for p, v in pc.items():
            phone_counts[p] = phone_counts.get(p, 0) + v
        skipped += skip
        print(f"   {haa_name}: {ok} мөр (2 эрх) · {skip} алгасав")

    wb.close()

    # ── Merge with existing data if --merge ──────────────────
    if merge:
        suffix = "" if month == 1 else str(month)
        existing_path = os.path.join(base_dir, f"data{suffix}.json")
        if os.path.exists(existing_path):
            with open(existing_path, encoding="utf-8") as f:
                existing = json.load(f)
            added = len(phone_counts)
            for p, v in existing.items():
                if p.startswith("_"):
                    continue
                phone_counts[p] = phone_counts.get(p, 0) + v
            # Also merge draw entries
            draw_path = os.path.join(base_dir, f"draw_data{suffix}.json")
            if os.path.exists(draw_path):
                with open(draw_path, encoding="utf-8") as f:
                    existing_draw = json.load(f)
                entries = existing_draw.get("entries", []) + entries
            print(f"   + Нэмэлт: {added} өвөрмөц дугаар одоогийн өгөгдөлд нэмэгдлээ")
        else:
            print(f"   ⚠️  {existing_path} олдсонгүй — шинэ файл үүсгэнэ")

    write_output(entries, phone_counts, skipped, month, base_dir)


def convert(excel_path: str, month: int = 1, merge: bool = False) -> None:
    suffix = "" if month == 1 else str(month)
    mode = "нэмэлт (merge)" if merge else "шинэ"
    print(f"\n📋 {month}-р сарын аян [{mode}] → data{suffix}.json / draw_data{suffix}.json\n")
    if not os.path.exists(excel_path):
        print(f"Файл олдсонгүй: {excel_path}")
        sys.exit(1)

    base_dir = os.path.dirname(os.path.abspath(excel_path))

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"\n❌  Excel файл унших боломжгүй: {e}")
        sys.exit(1)

    # ── Detect multi-sheet format ─────────────────────────
    # Matches: Data.xlsx (Зээл сунгалт/Хаасан зээл) AND daily files (sungalt/haasan/Сунгалт/etc.)
    sheet_names_lower = [s.strip().lower() for s in wb.sheetnames]
    has_sun = any("сунгалт" in s or "sungalt" in s for s in sheet_names_lower)
    has_haa = any("хаасан" in s or "haasan" in s for s in sheet_names_lower)
    if has_sun or has_haa:
        process_multi_sheet(wb, month, base_dir, merge=merge)
        return

    # ── Single-sheet formats ───────────────────────────────
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("❌ Хоосон файл.")
        sys.exit(1)

    headers = rows[0]
    fmt = detect_format(headers)
    fmt_labels = {
        'tpay': 'TPay экспорт (Төрөл|Регистр|Овог|Нэр|Утас) — Сунгалт=1эрх, Хаасан=2эрх',
        'new':  '5 багана (Нэр|Овог|Утас|Киоск|Огноо)',
        'old':  '2 багана (Утас|Тикет тоо)',
    }
    print(f"   Формат илэрлээ: {fmt_labels.get(fmt, fmt)}")

    # ── Build entries ──────────────────────────────────────
    entries = []
    phone_counts = {}
    skipped = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        if fmt == "tpay":
            töröl   = str(row[0]).strip() if row[0] else ""
            surname = str(row[2]).strip() if row[2] else ""
            name    = str(row[3]).strip() if row[3] else ""
            phone   = clean_phone(row[4])
            if töröl == "Сунгалт":
                tickets = 1
            elif töröl in ("Хаасан", "Хаах", "Хаасан зээл"):
                tickets = 2
            else:
                skipped += 1
                continue
        elif fmt == "new":
            name    = str(row[0]).strip() if row[0] else ""
            surname = str(row[1]).strip() if row[1] else ""
            phone   = clean_phone(row[2])
            kiosk   = str(row[3]).strip() if row[3] else ""
            loan_dt = fmt_date(row[4])
            tickets = 1
        else:
            phone   = clean_phone(row[0])
            name    = ""
            surname = ""
            kiosk   = ""
            loan_dt = ""
            try:
                tickets = int(float(str(row[1]))) if row[1] is not None else 1
            except Exception:
                tickets = 1

        if phone is None:
            skipped += 1
            continue

        if fmt == "tpay":
            for _ in range(tickets):
                entries.append({"phone": phone, "name": name, "surname": surname, "kiosk": "", "date": ""})
            phone_counts[phone] = phone_counts.get(phone, 0) + tickets
        elif fmt == "new":
            entries.append({"phone": phone, "name": name, "surname": surname, "kiosk": kiosk, "date": loan_dt})
            phone_counts[phone] = phone_counts.get(phone, 0) + 1
        else:
            for _ in range(max(tickets, 1)):
                entries.append({"phone": phone, "name": "", "surname": "", "kiosk": "", "date": ""})
            phone_counts[phone] = phone_counts.get(phone, 0) + max(tickets, 1)

    wb.close()
    write_output(entries, phone_counts, skipped, month, base_dir)


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    month_arg = 1
    merge_arg = False
    for i, a in enumerate(sys.argv[1:]):
        if a == "--month" and i + 2 < len(sys.argv):
            try:
                month_arg = int(sys.argv[i + 2])
            except ValueError:
                pass
        if a == "--merge":
            merge_arg = True
    if not args:
        print(__doc__)
        sys.exit(0)
    convert(args[0], month=month_arg, merge=merge_arg)
